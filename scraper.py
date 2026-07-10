"""Scrape recipes from app.groei-maatje.nl and export them to an Excel file.

The app is a Next.js application. Recipes are fetched through "server actions":
POST requests to the page URL with a `next-action` header identifying the
function to call and a JSON-array body with its arguments. Two actions are used:

  listRecipesForClientAction  -> the full recipe library (id, name, slug, ...)
  getRecipeForClientAction    -> full details for one recipe id

Authentication is the `__Secure-authjs.session-token` cookie, read from the
.env file next to this script (GM_SESSION_TOKEN). Grab a fresh token from your
browser DevTools (Application > Cookies) when the session expires.

Usage:
  python scraper.py                          # all recipes -> recipes.xlsx
  python scraper.py --out mijn_recepten.xlsx
  python scraper.py --ids cmpl09xca01yubjuhtifx2484 ...   # only specific ids
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_URL = "https://app.groei-maatje.nl/nutrition"
LIST_ACTION = "00dc5937d1b5daccd3a49856e2b91034ac975701e9"   # listRecipesForClientAction
RECIPE_ACTION = "40ba1edd6e61fcedccd3cedc7a05d8b4527afef27c"  # getRecipeForClientAction
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36"
)


def load_env() -> dict[str, str]:
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        sys.exit(f"Missing {env_path} - create it with GM_SESSION_TOKEN=<cookie value>")
    env = {}
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            env[key.strip()] = value.strip()
    if "GM_SESSION_TOKEN" not in env:
        sys.exit(f"{env_path} does not define GM_SESSION_TOKEN")
    return env


def call_server_action(session_token: str, action: str, args: list) -> object:
    """POST a Next.js server action and return the parsed result value."""
    request = urllib.request.Request(
        BASE_URL,
        data=json.dumps(args).encode(),
        headers={
            "Cookie": f"__Secure-authjs.session-token={session_token}",
            "User-Agent": USER_AGENT,
            "Accept": "text/x-component",
            "Content-Type": "text/plain;charset=UTF-8",
            "next-action": action,
            "Origin": "https://app.groei-maatje.nl",
            "Referer": BASE_URL,
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        body = response.read().decode("utf-8")
    # The response is React flight data: lines like `0:{...}` and `1:{...}`.
    # Line 1 holds the action's return value; it may itself contain newlines,
    # so parse from the start of the payload rather than splitting on lines.
    match = re.search(r"^1:", body, re.M)
    if not match:
        raise ValueError(f"No result row in server action response: {body[:200]!r}")
    value, _ = json.JSONDecoder().raw_decode(body[match.end():])
    return value


def fetch_with_retry(session_token: str, action: str, args: list, attempts: int = 3) -> object:
    for attempt in range(1, attempts + 1):
        try:
            return call_server_action(session_token, action, args)
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            if isinstance(exc, urllib.error.HTTPError) and exc.code in (401, 403):
                sys.exit("Session token rejected (401/403) - refresh GM_SESSION_TOKEN in .env")
            if attempt == attempts:
                raise
            time.sleep(2 * attempt)
    raise AssertionError("unreachable")


def ingredient_rows(recipe: dict) -> list[dict]:
    """Flatten a recipe's ingredients; base* fields in the API are per serving."""
    servings = recipe.get("servings") or 1
    rows = []
    for ing in recipe.get("ingredients", []):
        grams_per_serving = ing.get("baselineGrams") or 0
        rows.append(
            {
                "name": ing.get("foodItemName"),
                "brand": ing.get("brand"),
                "display_amount": ing.get("displayAmount"),
                "grams_total": round(grams_per_serving * servings, 1),
                "grams_per_serving": grams_per_serving,
                "unit": ing.get("baselineUnit"),
                "kcal": ing.get("baseKcal") or 0,
                "protein": ing.get("baseProtein") or 0,
                "carbs": ing.get("baseCarbs") or 0,
                "fat": ing.get("baseFat") or 0,
                "optional": ing.get("optional", False),
                "seasoning": ing.get("isSeasoning", False),
            }
        )
    return rows


def style_sheet(sheet, widths: list[int]) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_excel(recipes: list[dict], out_path: Path) -> None:
    workbook = Workbook()

    recipes_sheet = workbook.active
    recipes_sheet.title = "Recipes"
    recipes_sheet.append(
        [
            "Recipe", "Description", "Servings", "Serving description",
            "Prep (min)", "Cook (min)", "Kcal/serving", "Protein/serving (g)",
            "Carbs/serving (g)", "Fat/serving (g)", "Ingredients", "Steps",
            "Id", "Slug", "Cover image",
        ]
    )
    ingredients_sheet = workbook.create_sheet("Ingredients")
    ingredients_sheet.append(
        [
            "Recipe", "Ingredient", "Brand", "Amount", "Total grams",
            "Grams/serving", "Unit", "Kcal/serving", "Protein/serving (g)",
            "Carbs/serving (g)", "Fat/serving (g)", "Optional", "Seasoning",
            "Recipe id",
        ]
    )
    instructions_sheet = workbook.create_sheet("Instructions")
    instructions_sheet.append(["Recipe", "Step", "Instruction", "Recipe id"])

    for recipe in recipes:
        ingredients = ingredient_rows(recipe)
        recipes_sheet.append(
            [
                recipe.get("name"),
                recipe.get("description"),
                recipe.get("servings"),
                recipe.get("servingDescription"),
                recipe.get("prepTimeMin"),
                recipe.get("cookTimeMin"),
                round(sum(i["kcal"] for i in ingredients), 1),
                round(sum(i["protein"] for i in ingredients), 1),
                round(sum(i["carbs"] for i in ingredients), 1),
                round(sum(i["fat"] for i in ingredients), 1),
                len(ingredients),
                len(recipe.get("instructions", [])),
                recipe.get("id"),
                recipe.get("slug"),
                recipe.get("coverImageUrl"),
            ]
        )
        for ing in ingredients:
            ingredients_sheet.append(
                [
                    recipe.get("name"), ing["name"], ing["brand"],
                    ing["display_amount"], ing["grams_total"],
                    ing["grams_per_serving"], ing["unit"], ing["kcal"],
                    ing["protein"], ing["carbs"], ing["fat"],
                    "yes" if ing["optional"] else "no",
                    "yes" if ing["seasoning"] else "no",
                    recipe.get("id"),
                ]
            )
        for step_number, text in enumerate(recipe.get("instructions", []), start=1):
            instructions_sheet.append([recipe.get("name"), step_number, text, recipe.get("id")])

    style_sheet(recipes_sheet, [42, 60, 9, 24, 10, 10, 12, 16, 16, 14, 11, 7, 27, 40, 50])
    style_sheet(ingredients_sheet, [42, 28, 16, 22, 11, 13, 6, 12, 16, 16, 14, 8, 9, 27])
    style_sheet(instructions_sheet, [42, 6, 110, 27])
    for row in instructions_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export groei-maatje recipes to Excel")
    parser.add_argument("--out", default="recipes.xlsx", help="output Excel file (default: recipes.xlsx)")
    parser.add_argument("--ids", nargs="*", help="scrape only these recipe ids (default: whole library)")
    parser.add_argument("--delay", type=float, default=0.3, help="seconds between requests (default: 0.3)")
    options = parser.parse_args()

    token = load_env()["GM_SESSION_TOKEN"]

    if options.ids:
        targets = [{"id": recipe_id} for recipe_id in options.ids]
        print(f"Scraping {len(targets)} recipe(s) by id")
    else:
        print("Fetching recipe library ...")
        library = fetch_with_retry(token, LIST_ACTION, [])
        if not isinstance(library, list):
            sys.exit(f"Unexpected library response: {str(library)[:200]}")
        targets = library
        print(f"Library contains {len(targets)} recipes")

    recipes, failed = [], []
    for index, entry in enumerate(targets, start=1):
        recipe_id = entry["id"]
        label = entry.get("name", recipe_id)
        try:
            result = fetch_with_retry(token, RECIPE_ACTION, [recipe_id])
        except Exception as exc:
            print(f"  [{index}/{len(targets)}] FAILED {label}: {exc}")
            failed.append(recipe_id)
            continue
        if isinstance(result, dict) and result.get("ok") and result.get("recipe"):
            recipes.append(result["recipe"])
            print(f"  [{index}/{len(targets)}] {result['recipe']['name']}")
        else:
            print(f"  [{index}/{len(targets)}] SKIPPED {label}: {str(result)[:120]}")
            failed.append(recipe_id)
        time.sleep(options.delay)

    if not recipes:
        sys.exit("No recipes scraped - nothing to write")

    out_path = Path(options.out)
    write_excel(recipes, out_path)
    print(f"\nWrote {len(recipes)} recipes to {out_path.resolve()}")
    if failed:
        print(f"Failed ids ({len(failed)}): {', '.join(failed)}")


if __name__ == "__main__":
    main()
